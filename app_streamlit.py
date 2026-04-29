from __future__ import annotations

import pandas as pd
import streamlit as st
from datetime import date, timedelta, datetime
from io import BytesIO
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from scheduler_core import SolveConfig, solve_schedule_many, make_headers, NAME_MAP

DISPLAY_ORDER = [
    "최영철",
    "문승환",
    "라영일",
    "이병욱",
    "김동명",
    "홍진우",
    "김다영",
    "강예지",
    "이은주",
]

try:
    import holidays  # type: ignore
except Exception:
    holidays = None

try:
    from holidayskr import year_holidays  # type: ignore
except Exception:
    year_holidays = None

st.set_page_config(page_title="원무과 근무표 자동화", layout="wide")
st.title("🗓️ 원무과 근무표 자동 생성")
st.caption("표에서 전달 마지막 2일 + 이번 달 일정을 직접 입력한 뒤 생성합니다.")

today = date.today()
ALL_NAMES = list(NAME_MAP.values())
SUPPORT_NAME = NAME_MAP.get("S", "지원")

CHOICES_CURRENT = [".", "-", "휴", "D", "E", "N", "pD", "pE", "pN"]
CHOICES_PREV = [".", "-", "휴", "D", "E", "N"]


def days_in_month(y: int, m: int) -> int:
    import calendar
    return calendar.monthrange(y, m)[1]


def compute_stats(df: pd.DataFrame) -> pd.DataFrame:
    stats = pd.DataFrame(index=df.index)
    stats["D"] = (df == "D").sum(axis=1)
    stats["E"] = (df == "E").sum(axis=1)
    stats["N"] = (df == "N").sum(axis=1)
    stats["휴무"] = (df == "-").sum(axis=1)
    stats["휴가"] = (df == "휴").sum(axis=1)
    stats["총근무"] = stats["D"] + stats["E"] + stats["N"]
    stats["총일수"] = stats["총근무"] + stats["휴무"] + stats["휴가"]
    return stats


def _to_date_key(x) -> date | None:
    if x is None:
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, str):
        s = x.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        try:
            return date.fromisoformat(s)
        except Exception:
            return None
    try:
        if hasattr(x, "date"):
            dx = x.date()
            if isinstance(dx, date):
                return dx
    except Exception:
        pass
    return None


def get_kr_holidays_map(y: int) -> dict[date, str]:
    out: dict[date, str] = {}

    if year_holidays is not None:
        try:
            items = year_holidays(str(y))
            for d, name in items:
                dk = _to_date_key(d)
                if dk is not None:
                    out[dk] = str(name)
            if out:
                return out
        except Exception:
            pass

    if holidays is not None:
        try:
            hk = holidays.KR(years=y, observed=True)
            for d in hk:
                dk = _to_date_key(d)
                if dk is not None:
                    out[dk] = str(hk.get(d))
            if out:
                return out
        except TypeError:
            try:
                hk = holidays.KR(years=y)
                for d in hk:
                    dk = _to_date_key(d)
                    if dk is not None:
                        out[dk] = str(hk.get(d))
                if out:
                    return out
            except Exception:
                pass
        except Exception:
            pass

    return out


def weekday_kor(dt: date) -> str:
    kor_week = ["월", "화", "수", "목", "금", "토", "일"]
    return kor_week[dt.weekday()]


def build_day_label(dt: date, holiday_map: Optional[dict[date, str]] = None) -> str:
    base = f"{dt.month:02d}/{dt.day:02d}({weekday_kor(dt)})"
    if holiday_map and dt in holiday_map:
        base = f"{base}·{holiday_map[dt]}"
    if dt.weekday() >= 5 or (holiday_map and dt in holiday_map):
        base = f"🔴 {base}"
    return base


def normalize_prev_value(v: str) -> str:
    if v in [".", "휴", "", None]:
        return "-"
    if v in {"D", "E", "N", "-"}:
        return v
    return "-"


def normalize_current_value(v: str) -> str:
    if v in CHOICES_CURRENT:
        return v
    return "."


def build_editor_grid(year: int, month: int, names: List[str]) -> Tuple[pd.DataFrame, List[str], List[str], List[str]]:
    first_day = date(year, month, 1)
    prev2 = first_day - timedelta(days=2)
    prev1 = first_day - timedelta(days=1)

    holiday_map = get_kr_holidays_map(year)
    prev_labels = [build_day_label(prev2), build_day_label(prev1)]
    current_labels = [build_day_label(date(year, month, d), holiday_map) for d in range(1, days_in_month(year, month) + 1)]
    all_labels = prev_labels + current_labels

    df = pd.DataFrame(".", index=names, columns=all_labels)
    return df, prev_labels, current_labels, all_labels


def parse_editor_to_inputs(
    edited: pd.DataFrame,
    year: int,
    month: int,
    prev_labels: List[str],
    current_labels: List[str],
) -> Tuple[
    Dict[str, List[str]],
    Dict[str, Dict[str, str]],
    Dict[str, List[str]],
    Dict[str, Dict[str, List[str]]],
    Dict[str, List[str]],
]:
    prev_month_shifts: Dict[str, List[str]] = {}
    fixed_shifts: Dict[str, Dict[str, str]] = {}
    forced_off: Dict[str, List[str]] = {}
    preferred_shifts: Dict[str, Dict[str, List[str]]] = {}
    vacation_days: Dict[str, List[str]] = {}

    for name in edited.index:
        prev2 = normalize_prev_value(str(edited.loc[name, prev_labels[0]]).strip())
        prev1 = normalize_prev_value(str(edited.loc[name, prev_labels[1]]).strip())
        prev_month_shifts[name] = [prev2, prev1]

        fixed_map: Dict[str, str] = {}
        off_list: List[str] = []
        pref_map: Dict[str, List[str]] = []
        pref_dict: Dict[str, List[str]] = {}

        for d in range(1, days_in_month(year, month) + 1):
            col = current_labels[d - 1]
            iso = date(year, month, d).isoformat()
            v = normalize_current_value(str(edited.loc[name, col]).strip())

            if v in {"D", "E", "N", "-"}:
                fixed_map[iso] = v
            elif v == "휴":
                off_list.append(iso)
                vacation_days.setdefault(name, []).append(iso)
            elif v == "pD":
                pref_dict[iso] = ["D"]
            elif v == "pE":
                pref_dict[iso] = ["E"]
            elif v == "pN":
                pref_dict[iso] = ["N"]

        if fixed_map:
            fixed_shifts[name] = fixed_map
        if off_list:
            forced_off[name] = off_list
        if pref_dict:
            preferred_shifts[name] = pref_dict

    return prev_month_shifts, fixed_shifts, forced_off, preferred_shifts, vacation_days


def build_excel_like_request(out_df: pd.DataFrame, stats_df: pd.DataFrame, meta: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "근무표"

    year = meta["year"]
    month = meta["month"]
    start_date: date = meta["start_date"]
    num_days = meta["num_days"]
    red_dates = set(meta.get("red_dates", []))

    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="EDEDED")
    fill_red_bg = PatternFill("solid", fgColor="FFF2CC")
    fill_vac = PatternFill("solid", fgColor="D9E1F2")
    fill_support = PatternFill("solid", fgColor="FCE4D6")
    fill_summary = PatternFill("solid", fgColor="F7F7F7")

    font_header = Font(bold=True, size=11)
    font_title = Font(bold=True, size=16)
    font_cell = Font(bold=True, size=11)
    font_red = Font(bold=True, color="C00000")

    start_col = 2
    end_col = start_col + num_days - 1

    title = f"{year}년 {month:02d}월 원무팀 응급계 근무표"
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    tcell = ws.cell(1, start_col, title)
    tcell.font = font_title
    tcell.alignment = align_c

    ws.cell(1, end_col + 2, "월별지정off일수").font = Font(bold=True)
    ws.cell(1, end_col + 2).alignment = align_c
    ws.cell(1, end_col + 3, meta.get("min_off", "")).alignment = align_c

    ws.cell(2, 1, f"{month}월").font = Font(bold=True, size=12)
    ws.cell(2, 1).alignment = align_c

    kor_week = ["월", "화", "수", "목", "금", "토", "일"]
    dates = [start_date + timedelta(days=i) for i in range(num_days)]

    for i in range(num_days):
        c = start_col + i
        dt = dates[i]
        ws.cell(2, c, dt.day).alignment = align_c
        ws.cell(3, c, kor_week[dt.weekday()]).alignment = align_c

        for r in (2, 3):
            cell = ws.cell(r, c)
            cell.font = font_header
            cell.border = border
            cell.fill = fill_header
            cell.alignment = align_c

        if dt in red_dates:
            ws.cell(2, c).fill = fill_red_bg
            ws.cell(3, c).fill = fill_red_bg
            ws.cell(2, c).font = font_red
            ws.cell(3, c).font = font_red

    sum_headers = ["D", "E", "N", "휴무", "휴가", "총근무", "총일수"]
    for j, h in enumerate(sum_headers):
        col = end_col + 2 + j
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        cell = ws.cell(2, col, h)
        cell.font = font_header
        cell.alignment = align_c
        cell.fill = fill_header
        cell.border = border

    ws.column_dimensions["A"].width = 11
    for i in range(num_days):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 3.2
    for j in range(len(sum_headers)):
        ws.column_dimensions[get_column_letter(end_col + 2 + j)].width = 6

    base_row = 4
    for r, name in enumerate(out_df.index):
        row = base_row + r
        name_cell = ws.cell(row, 1, name)
        name_cell.alignment = align_l
        name_cell.border = border
        name_cell.font = Font(bold=True)

        for i in range(num_days):
            c = start_col + i
            dt = dates[i]
            val = str(out_df.iloc[r, i])
            cell = ws.cell(row, c, val)
            cell.alignment = align_c
            cell.font = font_cell
            cell.border = border
            if dt in red_dates:
                cell.fill = fill_red_bg
            if val == "휴":
                cell.fill = fill_vac
            if name == SUPPORT_NAME and val != "휴":
                cell.fill = fill_support

        s = stats_df.loc[name]
        values = [
            int(s["D"]),
            int(s["E"]),
            int(s["N"]),
            int(s["휴무"]),
            int(s["휴가"]),
            int(s["총근무"]),
            int(s["총일수"]),
        ]
        for j, v in enumerate(values):
            col = end_col + 2 + j
            cell = ws.cell(row, col, v)
            cell.alignment = align_c
            cell.border = border

    last_staff_row = base_row + len(out_df.index) - 1
    summary_start = last_staff_row + 2

    daily_D = [(out_df.iloc[:, i] == "D").sum() for i in range(num_days)]
    daily_E = [(out_df.iloc[:, i] == "E").sum() for i in range(num_days)]
    daily_N = [(out_df.iloc[:, i] == "N").sum() for i in range(num_days)]
    daily_V = [(out_df.iloc[:, i] == "휴").sum() for i in range(num_days)]

    rows = [
        ("D(배정)", daily_D),
        ("E(배정)", daily_E),
        ("N(배정)", daily_N),
        ("휴가(건수)", daily_V),
        ("N(필요)", meta.get("n_req_total", [""] * num_days)),
    ]

    for k, (label, values) in enumerate(rows):
        rr = summary_start + k
        ws.cell(rr, 1, label).font = Font(bold=True)
        ws.cell(rr, 1).alignment = align_l
        ws.cell(rr, 1).border = border
        for i in range(num_days):
            c = start_col + i
            dt = dates[i]
            cell = ws.cell(rr, c, values[i] if i < len(values) else "")
            cell.alignment = align_c
            cell.border = border
            cell.fill = fill_summary
            if dt in red_dates:
                cell.fill = fill_red_bg

    ws.freeze_panes = ws["B4"]

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


with st.sidebar:
    st.header("월 설정")
    year = st.number_input("연도", min_value=2020, max_value=2100, value=int(today.year), step=1)
    month = st.number_input("월", min_value=1, max_value=12, value=int(today.month), step=1)

    st.divider()
    st.header("일일 필요 인원")
    req_d_total = st.number_input("D(아침) / 일", min_value=0, max_value=10, value=2, step=1)
    req_e_total = st.number_input("E(저녁) / 일", min_value=0, max_value=10, value=2, step=1)
    use_min_off_minus_1 = st.checkbox("최소 휴무(빨간날) -1 허용", value=True)

    st.divider()
    st.header("사수/부사수 설정")
    base_pool = [n for n in ALL_NAMES if n != SUPPORT_NAME]
    mentors = st.multiselect("사수급(멘토)", options=base_pool, default=[])

    allowed_pool = [n for n in base_pool if n not in mentors]
    newbies_night_allowed = st.multiselect("부사수(나이트 허용)", options=allowed_pool, default=[])

    forbidden_pool = [n for n in allowed_pool if n not in newbies_night_allowed]
    newbies_night_forbidden = st.multiselect("부사수(나이트 불허용)", options=forbidden_pool, default=[])

    st.divider()
    st.header("지원(비상 인력)")
    use_support = st.checkbox("지원 사용", value=True)
    support_max_work = st.slider("지원 최대 투입(월)", min_value=0, max_value=31, value=6, step=1)

    st.divider()
    st.header("여러 안 생성")
    plan_k = st.slider("안 개수(k)", min_value=1, max_value=10, value=3, step=1)
    plan_max_gap = st.number_input("최적해 대비 허용 격차(max_gap)", min_value=0, max_value=1_000_000, value=30000, step=5000)
    plan_time = st.slider("안 하나당 탐색 시간(초)", min_value=5, max_value=120, value=30, step=5)

    st.divider()
    st.markdown("### 표 입력 규칙")
    st.write("전달 마지막 2일: `- / D / E / N / .`")
    st.write("이번 달: `-` 휴무, `휴` 휴가, `D/E/N` 고정근무, `pD/pE/pN` 희망, `.` 무관")

dup = (set(mentors) & set(newbies_night_allowed)) | (set(mentors) & set(newbies_night_forbidden)) | (set(newbies_night_allowed) & set(newbies_night_forbidden))
if dup:
    st.error(f"❌ 사수/부사수 카테고리에 중복이 있습니다: {sorted(list(dup))}")
    st.stop()

year = int(year)
month = int(month)
num_days = days_in_month(year, month)
start_date = date(year, month, 1)
dates = [start_date + timedelta(days=i) for i in range(num_days)]

kr_h_map = get_kr_holidays_map(year)
red_dates_ui = {d for d in dates if (d.weekday() >= 5 or d in kr_h_map)}

names_all = [n for n in DISPLAY_ORDER if n in ALL_NAMES and n != SUPPORT_NAME]

st.subheader(f"1) 입력 표 ({year}년 {month}월)")
st.write("맨 앞 2칸은 전달 마지막 2일입니다. 이 값은 월초 규칙 계산에만 반영되고 결과표/근무일수 계산에는 포함되지 않습니다.")

table_key = f"choice_table_{year}_{month}"
meta_key = f"choice_table_meta_{year}_{month}"

if table_key not in st.session_state:
    init_df, prev_labels, current_labels, all_labels = build_editor_grid(year, month, names_all)
    st.session_state[table_key] = init_df
    st.session_state[meta_key] = {
        "prev_labels": prev_labels,
        "current_labels": current_labels,
        "all_labels": all_labels,
    }
else:
    meta_saved = st.session_state.get(meta_key, {})
    saved_df = st.session_state[table_key]
    need_reset = False

    if not meta_saved:
        need_reset = True
    else:
        new_df, prev_labels, current_labels, all_labels = build_editor_grid(year, month, names_all)
        if list(saved_df.columns) != all_labels or list(saved_df.index) != names_all:
            need_reset = True

    if need_reset:
        init_df, prev_labels, current_labels, all_labels = build_editor_grid(year, month, names_all)
        st.session_state[table_key] = init_df
        st.session_state[meta_key] = {
            "prev_labels": prev_labels,
            "current_labels": current_labels,
            "all_labels": all_labels,
        }

prev_labels = st.session_state[meta_key]["prev_labels"]
current_labels = st.session_state[meta_key]["current_labels"]
all_labels = st.session_state[meta_key]["all_labels"]

col_a, col_b = st.columns([1, 1])
with col_a:
    if st.button("🧹 이번 달 입력표 초기화"):
        init_df, prev_labels, current_labels, all_labels = build_editor_grid(year, month, names_all)
        st.session_state[table_key] = init_df
        st.session_state[meta_key] = {
            "prev_labels": prev_labels,
            "current_labels": current_labels,
            "all_labels": all_labels,
        }
        st.rerun()
with col_b:
    st.caption("초기화는 현재 선택한 연/월 입력표만 초기화됩니다.")

column_config = {}
for col in all_labels:
    if col in prev_labels:
        column_config[col] = st.column_config.SelectboxColumn(col, options=CHOICES_PREV, required=True)
    else:
        column_config[col] = st.column_config.SelectboxColumn(col, options=CHOICES_CURRENT, required=True)

edited = st.data_editor(
    st.session_state[table_key],
    use_container_width=True,
    num_rows="fixed",
    column_config=column_config,
    key=f"editor_{table_key}",
)
st.session_state[table_key] = edited

prev_month_shifts, fixed_shifts, forced_off, preferred_shifts, vacation_days = parse_editor_to_inputs(
    edited=edited,
    year=year,
    month=month,
    prev_labels=prev_labels,
    current_labels=current_labels,
)

with st.expander("🔎 입력 확인", expanded=False):
    st.json({
        "prev_month_shifts": prev_month_shifts,
        "fixed_shifts": fixed_shifts,
        "forced_off": forced_off,
        "vacation_days": vacation_days,
        "preferred_shifts": preferred_shifts,
        "mentors": mentors,
        "newbies_night_allowed": newbies_night_allowed,
        "newbies_night_forbidden": newbies_night_forbidden,
        "use_support": use_support,
        "support_max_work": support_max_work,
        "holiday_count_detected": len(kr_h_map),
        "holiday_sample": sorted([(d.isoformat(), kr_h_map[d]) for d in kr_h_map if d.month == month])[:10],
        "k": int(plan_k),
        "max_gap": int(plan_max_gap),
        "time_per_solve": int(plan_time),
    })

with st.expander("📋 이번 달 입력만 보기", expanded=False):
    st.dataframe(edited[current_labels], use_container_width=True)

if year_holidays is None and holidays is None:
    st.warning("공휴일/대체공휴일 탐지를 위해 `pip install holidayskr`(권장) 또는 `pip install holidays`를 설치하세요.")

st.subheader("2) 근무표 생성")
run = st.button("🧩 근무표 생성", type="primary")

cache_key = f"sol_cache_{year}_{month}"

if run:
    cfg = SolveConfig(
        year=year,
        month=month,
        req_d_total=int(req_d_total),
        req_e_total=int(req_e_total),
        use_min_off_minus_1=bool(use_min_off_minus_1),
        prev_month_shifts=prev_month_shifts if prev_month_shifts else None,
        fixed_shifts=fixed_shifts if fixed_shifts else None,
        forced_off=forced_off if forced_off else None,
        preferred_shifts=preferred_shifts if preferred_shifts else None,
        mentors=mentors if mentors else None,
        newbies_night_allowed=newbies_night_allowed if newbies_night_allowed else None,
        newbies_night_forbidden=newbies_night_forbidden if newbies_night_forbidden else None,
        use_support=bool(use_support),
        support_max_work=int(support_max_work),
        support_name=SUPPORT_NAME,
    )

    with st.spinner("근무표 계산 중..."):
        try:
            sols, mb = solve_schedule_many(
                cfg,
                k=int(plan_k),
                max_gap=int(plan_max_gap),
                time_per_solve=float(plan_time),
            )
            st.session_state[cache_key] = {
                "solutions": sols or [],
                "meta_base": mb or {},
                "error": None,
                "cfg_sig": (
                    year, month, int(req_d_total), int(req_e_total),
                    bool(use_min_off_minus_1), bool(use_support), int(support_max_work),
                    int(plan_k), int(plan_max_gap), int(plan_time),
                    tuple(sorted((k, tuple(v)) for k, v in prev_month_shifts.items())),
                    tuple(sorted((k, tuple(sorted(v.items()))) for k, v in fixed_shifts.items())),
                    tuple(sorted((k, tuple(v)) for k, v in forced_off.items())),
                    tuple(sorted((k, tuple(sorted((d, tuple(c)) for d, c in v.items()))) for k, v in preferred_shifts.items())),
                    tuple(sorted(mentors)),
                    tuple(sorted(newbies_night_allowed)),
                    tuple(sorted(newbies_night_forbidden)),
                ),
            }
        except Exception as e:
            st.session_state[cache_key] = {
                "solutions": [],
                "meta_base": {},
                "error": repr(e),
                "cfg_sig": None,
            }

payload = st.session_state.get(cache_key)

if payload is None:
    st.info("👈 전달 마지막 2일과 이번 달 표를 입력한 뒤 **근무표 생성**을 누르세요.")
    st.stop()

solutions = payload.get("solutions", []) or []
meta_base = payload.get("meta_base", {}) or {}
cached_err = payload.get("error")

if not solutions:
    st.error("❌ 해를 찾지 못했습니다. (하드 규칙 충돌 가능)")
    if cached_err:
        st.write(cached_err)
    st.stop()

labels = [f"{i+1}안 (obj={obj})" for i, (obj, _) in enumerate(solutions)]
choice = st.selectbox(
    "근무표 선택",
    options=list(range(len(solutions))),
    format_func=lambda i: labels[i],
    key=f"plan_choice_{cache_key}",
)

objv, schedule = solutions[choice]
meta = dict(meta_base)
meta["objective"] = int(objv)
meta["year"] = year
meta["month"] = month
meta["start_date"] = start_date
meta["num_days"] = num_days
meta["red_dates"] = sorted(list(red_dates_ui))
meta["n_req_total"] = [""] * num_days

headers = make_headers(meta)
out_df = pd.DataFrame(schedule).T
out_df.columns = headers

result_order = [n for n in DISPLAY_ORDER if n in out_df.index]
remaining = [n for n in out_df.index if n not in result_order]
out_df = out_df.loc[result_order + remaining]

if vacation_days and len(headers) == num_days:
    current_iso = [date(year, month, d).isoformat() for d in range(1, num_days + 1)]
    for person, iso_list in vacation_days.items():
        if person not in out_df.index:
            continue
        iso_set = set(iso_list)
        for i, _ in enumerate(headers):
            if current_iso[i] in iso_set:
                out_df.loc[person, headers[i]] = "휴"

stats_df = compute_stats(out_df)
final_df = pd.concat([out_df, stats_df], axis=1)

num_days_cols = list(final_df.columns[:num_days])
red_col_mask = [(dates[i] in red_dates_ui) for i in range(num_days)]

def _highlight_red_cols(_):
    styles = pd.DataFrame("", index=final_df.index, columns=final_df.columns)
    for i, is_red in enumerate(red_col_mask):
        if is_red:
            styles[num_days_cols[i]] = "background-color: #FFF2CC"
    return styles

st.success(f"✅ 근무표 생성 완료: {labels[choice]} (총 {len(solutions)}개)")
st.dataframe(final_df.style.apply(_highlight_red_cols, axis=None), use_container_width=True)

xlsx_bytes = build_excel_like_request(out_df, stats_df, meta)
st.download_button(
    "⬇ 엑셀(.xlsx) 다운로드",
    data=xlsx_bytes,
    file_name=f"schedule_{meta['year']}_{meta['month']:02d}_{choice+1}an.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)